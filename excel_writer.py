"""
Excel反映モジュール

抽出した明細データを、既存の集計用.xlsxに新シートとして追加する。
分類ルール（社保/生産課/材料費）も自動適用する。

邸数によって動的に行を挿入することで18邸を超える場合でも対応する。
基準レイアウト:
    5-22: 明細データ行 (18行)
    23:   スペーサー（空）
    24:   合計行 (SUM式)
    29-31: 班長集計 (SUMIF)
    29-35: 赤枠エリア
    37-44: 振込金額照合
邸数が18を超える場合、行23の前に行を挿入することで範囲を拡張する。
"""

import datetime
import re
from typing import Optional
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import FormulaRule, CellIsRule, Rule
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
from collections import defaultdict


DEFAULT_DATA_ROWS = 18  # テンプレ既定の明細行数(5-22)
MAX_TEI = 50             # これ以上は明示的にエラー


def classify_and_aggregate(rows: list[dict]) -> list[dict]:
    """明細行を邸ごとに集計し、D/E/F/G列に振り分ける。"""
    by_tei = defaultdict(lambda: {
        "邸名": "",
        "契約NO": set(),
        "工事名称": set(),
        "D_items": [],
        "E": 0,
        "F": 0,
        "G_items": [],
    })

    for row in rows:
        tei = row["邸名"]
        amount_raw = row["税抜金額"]
        koushu = row["工種"]
        bikou = row.get("備考", "")

        if not tei or tei in ("計", "合計") or "消費税" in tei or "対象外" in tei:
            print(f"  [skip] 集計行: 邸名={tei}")
            continue

        try:
            amount = int(round(float(amount_raw)))
        except (TypeError, ValueError):
            print(f"  [skip] 金額パース失敗: 邸={tei} 金額={amount_raw!r}")
            continue

        agg = by_tei[tei]
        agg["邸名"] = tei
        agg["契約NO"].add(row.get("契約NO", ""))

        base_name = _extract_koji_base(koushu)
        if base_name:
            agg["工事名称"].add(base_name)

        # 緩和マッチ(v1.0.98〜): 「中口応援分」等のバリエーションも生産課扱い
        # pdfplumberが「生産課中口分」を「生産課」と「中口分」に分割するケース対応 +
        # 「中口応援分」など中口プレフィックスを生産課判定に含める(山本さん指示)
        is_seisanka = ("生産課" in bikou) or ("中口" in bikou)
        is_shaho = "社保" in koushu

        if amount >= 0:
            agg["D_items"].append(amount)
        else:
            abs_amount = abs(amount)
            if is_seisanka and is_shaho:
                agg["E"] += abs_amount
                print(f"  [classify] E(社保): 邸={tei} 金額={amount} 工種={koushu} 備考={bikou!r}")
            elif is_seisanka:
                agg["F"] += abs_amount
                print(f"  [classify] F(生産課): 邸={tei} 金額={amount} 工種={koushu} 備考={bikou!r}")
            else:
                agg["G_items"].append(abs_amount)
                print(f"  [classify] G(材料費): 邸={tei} 金額={amount} 工種={koushu} 備考={bikou!r}")

    result = []
    for tei, agg in by_tei.items():
        koji_names = list(agg["工事名称"])
        koji_label = "・".join(sorted(set(koji_names))) if koji_names else ""
        result.append({
            "邸名": tei,
            "契約NO": list(agg["契約NO"])[0] if agg["契約NO"] else "",
            "工事名称": koji_label,
            "D_items": agg["D_items"],
            "E": agg["E"],
            "F": agg["F"],
            "G_items": agg["G_items"],
        })
    return result


def _extract_koji_base(koushu: str) -> Optional[str]:
    if "防水" in koushu:
        return "防水"
    if "柱脚" in koushu:
        return "柱脚"
    return None


def write_to_template(
    template_path: str,
    output_path: str,
    sheet_name: str,
    aggregated: list[dict],
    furikomi_kingaku: Optional[int] = None,
    pdf_koujidai_zeikomi: Optional[int] = None,
    pdf_sousai_zeikomi: Optional[int] = None,
    payment_date: Optional[str] = None,
):
    """集計用テンプレートに書き込む。邸数に応じて動的に行挿入する。"""
    n_tei = len(aggregated)
    if n_tei < 1:
        raise ValueError("抽出された邸数が0です。PDFの内容を確認してください。")
    if n_tei > MAX_TEI:
        raise ValueError(f"邸数が{MAX_TEI}を超えています: {n_tei}邸")

    # 年次ファイル運用: output_path に既存ファイルがあればそれを、なければテンプレート
    import os as _os
    base_path = output_path if _os.path.exists(output_path) else template_path
    wb = load_workbook(base_path)

    # 年抽出: "2026年4月" → "2026"
    m_year = re.search(r'(\d{4})年', sheet_name)
    year_str = m_year.group(1) if m_year else ""

    # 賞与シート名+セル内テキストのプレースホルダ〇〇〇〇を年で置換
    # (シート名は前回ランで既に置換済みでも、セル内のB2タイトルなどは残ってるため毎回スキャン)
    if year_str:
        for sn in list(wb.sheetnames):
            if '〇〇〇〇' in sn:
                wb[sn].title = sn.replace('〇〇〇〇', year_str)
        for sn in wb.sheetnames:
            if '夏' in sn or '冬' in sn or '賞与' in sn:
                _replace_placeholder_in_cells(wb[sn], '〇〇〇〇', year_str)

    # sheet_name "2026年4月" → 月シート "4月" を探す（賞与シート参照を壊さないため月シートはリネームしない）
    m_month = re.search(r'(\d{1,2})月', sheet_name)
    target_sheet_name = f"{m_month.group(1)}月" if m_month else sheet_name
    if target_sheet_name not in wb.sheetnames:
        # 単一テンプレ時代の後方互換: 先頭シートを使ってリネーム
        ws = wb[wb.sheetnames[0]]
        ws.title = sheet_name
    else:
        ws = wb[target_sheet_name]
        # 月シートの名前は保持（"4月"のまま）

    # 全mergeを解除（read-only エラー回避）
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))

    # 合計行を常に 5+n_tei 行に統一（邸数に関わらずデータ行の直下）
    # 既存ファイル(年次運用)の上書き時に1行ずつドリフトするのを防ぐため、
    # 「テンプレ配置(spacer+sum)」と「圧縮済み配置(sumのみ)」の両方を検出して
    # 必要な差分だけ insert/delete する。
    desired_sum_row = 5 + n_tei
    existing_sum_row = _detect_existing_sum_row(ws)
    has_spacer = _detect_spacer(ws, existing_sum_row)

    if has_spacer:
        # テンプレ状態(またはspacer残存): spacer+余分データ行を消して desired_sum_row に揃える
        if desired_sum_row >= existing_sum_row:
            # n_tei >= テンプレデータ行数: spacer手前にデータ行を挿入(spacerと同数+1)→ spacer削除
            # n=19: insert 1 + delete 1 = net 0 (spacerをデータ行に変換するイメージ)
            # n=20: insert 2 + delete 1 = net +1
            insert_count = desired_sum_row - existing_sum_row + 1
            ws.insert_rows(existing_sum_row - 1, amount=insert_count)
            _copy_data_format(ws, src_row=existing_sum_row - 2,
                              dst_rows=range(existing_sum_row - 1,
                                             existing_sum_row - 1 + insert_count))
            ws.delete_rows(desired_sum_row, amount=1)  # 挿入でズレたspacerを削除
            print(f"  [insert+spacer変換] +{insert_count - 1}行 ({n_tei}邸, テンプレ状態)")
        else:
            # n_tei < テンプレデータ行数: 余分データ行+spacerを削除(sum行は残して上にシフトさせる)
            delete_count = existing_sum_row - desired_sum_row
            if delete_count > 0:
                ws.delete_rows(desired_sum_row, amount=delete_count)
                print(f"  [delete] -{delete_count}行 ({n_tei}邸, テンプレ状態)")
    else:
        # 圧縮済み状態(既存output): 既存sum_rowからdesiredへの差分のみ
        delta = desired_sum_row - existing_sum_row
        if delta > 0:
            ws.insert_rows(existing_sum_row, amount=delta)
            _copy_data_format(ws, src_row=existing_sum_row - 1,
                              dst_rows=range(existing_sum_row, existing_sum_row + delta))
            print(f"  [insert] +{delta}行 ({n_tei}邸, 圧縮状態から拡張)")
        elif delta < 0:
            ws.delete_rows(desired_sum_row, amount=-delta)
            print(f"  [delete] {delta}行 ({n_tei}邸, 圧縮状態から縮小)")
        else:
            print(f"  [no-op] レイアウト変更なし ({n_tei}邸, 既存と一致)")

    # 行番号ヘルパー（合計行は常に 5 + n_tei）
    data_last_row = 4 + n_tei
    sum_row = 5 + n_tei
    hancho_row_start = sum_row + 5      # 元24→29なので+5
    furikomi_start = sum_row + 13       # 元24→37なので+13
    extra = 0  # 後続のレガシーなオフセット計算を無効化

    # 赤枠再描画（合計行=5+n_tei 基準、元テンプレB29:J35 → B(sum_row+5):J(sum_row+11)）
    red_top = sum_row + 5
    red_bottom = sum_row + 11
    _draw_red_border(ws, top=red_top, bottom=red_bottom, left=2, right=10)

    # タイトル
    ws['C2'] = f'{sheet_name}　着工=受注　ベース'

    # K1: 更新日（今日の日付）
    today = datetime.date.today()
    ws['K1'] = f'{today.year}/{today.month}/{today.day} 更新'
    # K2: 支払日（PDFから抽出、無ければ空）
    if payment_date:
        ws['K2'] = f'支払日: {payment_date}'
    else:
        ws['K2'] = None

    # 旧レイアウトの注釈セル(K27:L27相当)をクリア（sum_row + 3）
    note_row = sum_row + 3
    ws.cell(row=note_row, column=11).value = None
    ws.cell(row=note_row, column=12).value = None

    # 過去の上書き運用で残った「ドリフト残骸」を掃除
    # (i) hancho_row_start 直上 (sum_row+3 〜 sum_row+4) の K/L 残骸
    for r in range(sum_row + 3, hancho_row_start):
        ws.cell(row=r, column=11).value = None
        ws.cell(row=r, column=12).value = None
    # (ii) ラベル配置(陸くん指定): 売上合計→C{sum_row}, 原材料経費合計→H{sum_row+1}, 利益→J{sum_row+1}
    # 元テンプレ位置(C{sum_row+1}, E{sum_row+1}) のラベル残骸はクリア(値+塗り)してから新位置に書き込む
    label_row = sum_row + 1
    _clear_label_with_fill(ws, label_row, 3, '売上合計')             # 旧 C{sum_row+1}
    _clear_label_with_fill(ws, label_row, 5, '原材料　経費　合計')   # 旧 E{sum_row+1}
    _restore_label_if_missing(ws, sum_row, 3, '売上合計')         # 新 C{sum_row}
    _restore_label_if_missing(ws, label_row, 8, '原材料　経費　合計')  # 新 H{sum_row+1}
    _restore_label_if_missing(ws, label_row, 10, '利益')          # J{sum_row+1} は据え置き
    # H{sum_row+1} の見栄え(陸くん指定): Meiryo 17 + 右寄せ + 赤塗り
    h_cell = ws.cell(row=label_row, column=8)
    h_cell.font = Font(name='Meiryo', size=17)
    h_cell.alignment = Alignment(horizontal='right', vertical='center')
    # 一つ上のセル(H{sum_row})と同じ塗り(陸くん指定: FFFDE9D9 薄ピーチ)
    h_cell.fill = PatternFill(start_color='FFFDE9D9', end_color='FFFDE9D9', fill_type='solid')

    # C列赤塗りクリア（全データ行）
    no_fill = PatternFill(fill_type=None)
    for r in range(5, data_last_row + 1):
        ws.cell(row=r, column=3).fill = no_fill

    # 明細書き込み
    _write_rows(ws, aggregated, data_last_row)

    # 合計行の数式を挿入後の範囲で書き換え
    _rewrite_sum_row(ws, sum_row, data_last_row)

    # 付帯数式（原材料経費合計・生産課支払）: sum_row + 1, sum_row + 2
    i_zairyo_row = sum_row + 1
    e_seisanka_row = sum_row + 2
    ws[f'I{i_zairyo_row}'] = f'=SUM(E5:I{data_last_row})'
    ws[f'E{e_seisanka_row}'] = f'=SUM(E{sum_row}:F{sum_row})'
    # 「生産課 支払 "数値」のセルはMeiryo 20なので E+F マージで表示幅を確保
    try:
        ws.merge_cells(start_row=e_seisanka_row, start_column=5, end_row=e_seisanka_row, end_column=6)
    except Exception:
        pass

    # 担当者別集計(SUMIF) - 挿入後の行/範囲
    # K列の班長名も明示的に書き込み(ドリフト済みファイルでも正しい位置に復元される)
    data_range_J = f"J5:J{data_last_row}"
    data_range_K = f"K5:K{data_last_row}"
    HANCHO_NAMES = ('山本', '熱田', '安保')
    for i, name in enumerate(HANCHO_NAMES):
        r = hancho_row_start + i
        ws.cell(row=r, column=11, value=name)
        ws.cell(row=r, column=12, value=f'=SUMIF({data_range_K},K{r},{data_range_J})')

    # 振込金額照合
    _write_furikomi_verification(
        ws, furikomi_kingaku, pdf_sousai_zeikomi,
        start_row=furikomi_start, sum_row=sum_row,
    )

    # 使いやすさ機能
    _add_usability_features(ws, data_last_row=data_last_row, furikomi_start=furikomi_start)

    try:
        wb.save(output_path)
    finally:
        wb.close()


def _replace_placeholder_in_cells(ws, placeholder: str, replacement: str):
    """シート全体のテキストセルから placeholder を replacement に置換。
    マージセルやスタイルは触らない(値の中の文字列のみ書き換え)。"""
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and placeholder in v:
                try:
                    cell.value = v.replace(placeholder, replacement)
                except (AttributeError, TypeError):
                    pass  # MergedCell の slave 等は無視


def _clear_if_label(ws, row: int, col: int, label: str):
    """指定セルが指定ラベル文字列だった場合のみクリア(他の値は触らない)。
    ラベル位置を移動する時に旧位置をクリアする用途。"""
    v = ws.cell(row=row, column=col).value
    if v == label:
        ws.cell(row=row, column=col).value = None


def _clear_label_with_fill(ws, row: int, col: int, label: str):
    """指定ラベルだったら値も塗り(背景色)も両方クリア。
    旧位置に色付きセルが残るのを防ぐ。"""
    cell = ws.cell(row=row, column=col)
    if cell.value == label:
        cell.value = None
    # ラベル位置だったセル(値が空 or 元ラベルだったケース)の色塗りを除去
    if cell.value is None:
        cell.fill = PatternFill(fill_type=None)


def _restore_label_if_missing(ws, row: int, col: int, label: str):
    """指定セルが空 or 数式 or 数値の場合、テキストラベルを復元する。
    既に正しいテキストが入っている場合は何もしない(書式が壊れない)。"""
    v = ws.cell(row=row, column=col).value
    if v == label:
        return
    if not isinstance(v, str) or v.startswith('='):
        ws.cell(row=row, column=col, value=label)


def _detect_existing_sum_row(ws) -> int:
    """既存シートの合計行(=SUM(D5:Dn) を含む行)を D列スキャンで検出。
    見つからない場合は素のテンプレ想定で 24 を返す。"""
    for r in range(5, 60):
        v = ws.cell(row=r, column=4).value
        if isinstance(v, str) and v.upper().startswith('=SUM(D5'):
            return r
    return 24


def _detect_spacer(ws, sum_row: int) -> bool:
    """合計行の直上が空(=spacer)かどうか判定。テンプレ状態 vs 圧縮済みの分岐用。"""
    above = ws.cell(row=sum_row - 1, column=4).value
    return above is None or above == ''


def _copy_data_format(ws, src_row: int, dst_rows):
    """データ行の書式を複製し、J列(粗利)/L列(粗利率)に行ごとの数式を埋め込む。"""
    src_height = ws.row_dimensions[src_row].height
    for new_r in dst_rows:
        ws.row_dimensions[new_r].height = src_height
        for col_idx in range(1, 15):
            src_cell = ws.cell(row=src_row, column=col_idx)
            dst_cell = ws.cell(row=new_r, column=col_idx)
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.border = copy(src_cell.border)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.number_format = src_cell.number_format
                dst_cell.alignment = copy(src_cell.alignment)
                dst_cell.protection = copy(src_cell.protection)
            if col_idx == 10:
                dst_cell.value = f'=ROUNDDOWN(D{new_r}-E{new_r}-F{new_r}-G{new_r}-H{new_r}-I{new_r},0)'
            elif col_idx == 12:
                dst_cell.value = f'=IFERROR(J{new_r}/D{new_r},"")'


def _draw_red_border(ws, top: int, bottom: int, left: int, right: int):
    """指定範囲の外周に赤枠(medium)を描画。内側エッジはクリア。"""
    red_side = Side(style='medium', color='FFC00000')
    no_side = Side(style=None)
    for row in range(top, bottom + 1):
        for col_idx in range(left, right + 1):
            is_top = row == top
            is_bottom = row == bottom
            is_left = col_idx == left
            is_right = col_idx == right
            if not (is_top or is_bottom or is_left or is_right):
                continue
            ws.cell(row=row, column=col_idx).border = Border(
                left=red_side if is_left else no_side,
                right=red_side if is_right else no_side,
                top=red_side if is_top else no_side,
                bottom=red_side if is_bottom else no_side,
            )


def _write_rows(ws, aggregated: list[dict], data_last_row: int):
    """明細データをワークシートに書き込む。余ったデータ行はクリア。"""
    for i, item in enumerate(aggregated):
        r = 5 + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=item["邸名"])
        ws.cell(row=r, column=3, value=item["工事名称"])
        d_formula = '=' + '+'.join(str(x) for x in item["D_items"]) if item["D_items"] else 0
        ws.cell(row=r, column=4, value=d_formula)
        ws.cell(row=r, column=5, value=item["E"] if item["E"] else None)
        ws.cell(row=r, column=6, value=item["F"] if item["F"] else None)
        if item["G_items"]:
            ws.cell(row=r, column=7, value='=' + '+'.join(str(x) for x in item["G_items"]))
        else:
            ws.cell(row=r, column=7, value=None)
        ws[f'H{r}'].value = None
        ws[f'I{r}'].value = None
        ws[f'K{r}'].value = None

    # 余った行をクリア（A列の行番号 + B-L列すべて）
    n = len(aggregated)
    for r in range(5 + n, data_last_row + 1):
        for col in range(1, 13):  # A..L
            ws.cell(row=r, column=col).value = None


def _rewrite_sum_row(ws, sum_row: int, data_last_row: int):
    """合計行の数式を新しい範囲で書き換える。"""
    # D〜I: SUM
    for col_letter in ['D', 'E', 'F', 'G', 'H', 'I']:
        ws[f'{col_letter}{sum_row}'] = f'=SUM({col_letter}5:{col_letter}{data_last_row})'
    # J: 粗利合計（プラス連結式で精度確保）
    j_terms = '+'.join(f'J{r}' for r in range(5, data_last_row + 1))
    ws[f'J{sum_row}'] = f'=ROUNDDOWN({j_terms},0)'
    # L: 粗利率
    ws[f'L{sum_row}'] = f'=IFERROR(J{sum_row}/D{sum_row},"")'


def _write_furikomi_verification(ws, furikomi, sousai, start_row: int, sum_row: int):
    """振込金額照合欄（税抜⇔税込の二重計算）"""
    r_header = start_row
    r_furikomi = start_row + 1
    r_sousai = start_row + 2
    r_zeikomi_total = start_row + 3
    r_zeinuki_calc = start_row + 4
    r_excel_total = start_row + 5
    r_sagaku = start_row + 6
    r_note = start_row + 7

    # 新位置クリア + ヘッダ直上のドリフト残骸も掃除
    # (上書き運用で B34/B35 等にヘッダ重複が出るのを防ぐ)
    scrub_top = max(sum_row + 8, r_header - 6)
    for r in range(scrub_top, r_note + 1):
        for c in range(2, 6):
            ws.cell(row=r, column=c).value = None

    ws.cell(row=r_header, column=2, value='【振込金額照合（税抜⇔税込の二重計算）】')
    ws.cell(row=r_header, column=2).font = copy(ws['C2'].font)

    ws.cell(row=r_furikomi, column=2, value='① 振込金額(税込)')
    ws.cell(row=r_furikomi, column=4, value=furikomi if furikomi is not None else None)
    ws.cell(row=r_sousai, column=2, value='② 税込相殺(PDF・手入力)')
    ws.cell(row=r_sousai, column=4, value=sousai if sousai is not None else None)
    ws.cell(row=r_zeikomi_total, column=2, value='③ 税込工事代計(① − ②)')
    ws.cell(row=r_zeikomi_total, column=4, value=f'=D{r_furikomi}-D{r_sousai}')
    # ③ 行のboldを解除（テンプレ由来で太字になってるため）
    b_font = ws.cell(row=r_zeikomi_total, column=2).font
    d_font = ws.cell(row=r_zeikomi_total, column=4).font
    ws.cell(row=r_zeikomi_total, column=2).font = Font(
        name=b_font.name, size=b_font.size or 17, bold=False, color=b_font.color
    )
    ws.cell(row=r_zeikomi_total, column=4).font = Font(
        name=d_font.name, size=d_font.size or 17, bold=False, color=d_font.color
    )
    ws.cell(row=r_zeinuki_calc, column=2, value='④ 税抜逆算(③ ÷ 1.1)')
    ws.cell(row=r_zeinuki_calc, column=4, value=f'=ROUND(D{r_zeikomi_total}/1.1,0)')
    ws.cell(row=r_excel_total, column=2, value=f'⑤ Excel税抜合計(J{sum_row})')
    ws.cell(row=r_excel_total, column=4, value=f'=J{sum_row}')
    ws.cell(row=r_sagaku, column=2, value='⑥ 差額(⑤ − ④)')
    ws.cell(row=r_sagaku, column=4, value=f'=D{r_excel_total}-D{r_zeinuki_calc}')
    ws.cell(row=r_note, column=2, value='※ ±数円→インボイス端数差(正常) / 大きな差→PDF読取エラーの可能性')

    for r in [r_furikomi, r_sousai, r_zeikomi_total, r_zeinuki_calc, r_excel_total, r_sagaku]:
        ws.cell(row=r, column=4).number_format = '#,##0;[Red]▲#,##0'

    # ⑥差額 行は他の振込金額照合行と同じサイズ(17)、太字にしない
    ws.cell(row=r_sagaku, column=2).font = Font(
        name=ws.cell(row=r_furikomi, column=2).font.name, size=17, bold=False
    )
    ws.cell(row=r_sagaku, column=4).font = Font(
        name=ws.cell(row=r_furikomi, column=4).font.name, size=17, bold=False
    )


def _add_usability_features(ws, data_last_row: int, furikomi_start: int):
    """プルダウン／条件付き書式／担当邸数カウント"""
    k_range = f"K5:K{data_last_row}"
    b_range = f"B5:B{data_last_row}"

    # 班長プルダウン
    dv = DataValidation(
        type='list', formula1='"山本,熱田,安保"', allow_blank=True,
        showErrorMessage=True, errorTitle='班長名エラー',
        error='山本 / 熱田 / 安保 から選んでください',
    )
    dv.add(k_range)
    ws.add_data_validation(dv)

    # 既存の条件付き書式をリセット（過去ルールが挿入で壊れている可能性があるため）
    ws.conditional_formatting = ConditionalFormattingList()

    # 差額赤/緑 (D{sagaku_row})
    sagaku_row = furikomi_start + 6  # ⑥ 差額 の行
    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    green_fill = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
    ws.conditional_formatting.add(
        f'D{sagaku_row}', FormulaRule(formula=[f'ABS(D{sagaku_row})>10'], fill=red_fill)
    )
    ws.conditional_formatting.add(
        f'D{sagaku_row}', FormulaRule(formula=[f'AND(ABS(D{sagaku_row})<=10,D{sagaku_row}<>"")'], fill=green_fill)
    )

    # 班長未入力黄色
    light_yellow = PatternFill(start_color='FFF9E6', end_color='FFF9E6', fill_type='solid')
    ws.conditional_formatting.add(
        k_range, FormulaRule(formula=[f'AND(K5="",B5<>"")'], fill=light_yellow)
    )

    # 班長名による配置と色分け
    _add_hancho_styling(ws, k_range)

    # L列(粗利率)右揃え統一
    for r in range(5, data_last_row + 1):
        ws[f'L{r}'].alignment = Alignment(horizontal='right', vertical='center')

    # 担当邸数カウント(N3:O9) - 固定位置（挿入された行に影響されない想定）
    ws['N3'] = '【担当邸数】'
    ws['N3'].font = copy(ws['C2'].font)
    ws['N4'] = '班長'
    ws['O4'] = '邸数'
    header_font = Font(name=ws['B5'].font.name, bold=True)
    ws['N4'].font = header_font
    ws['O4'].font = header_font
    ws['N5'] = '山本'
    ws['O5'] = f'=COUNTIF({k_range},N5)'
    ws['N6'] = '熱田'
    ws['O6'] = f'=COUNTIF({k_range},N6)'
    ws['N7'] = '安保'
    ws['O7'] = f'=COUNTIF({k_range},N7)'
    ws['N8'] = '未入力'
    ws['O8'] = f'=COUNTBLANK({k_range})-COUNTBLANK({b_range})'
    ws['N9'] = '合計'
    ws['O9'] = '=SUM(O5:O8)'
    ws.conditional_formatting.add('O8', CellIsRule(operator='greaterThan', formula=['0'], fill=light_yellow))
    ws.column_dimensions['N'].width = 14
    ws.column_dimensions['O'].width = 10
    for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col].width = 28


def _add_hancho_styling(ws, k_range: str):
    """班長名に応じて配置と色を自動変更する条件付き書式。"""
    styles = [
        ("山本", "left", "FF006100"),
        ("熱田", "center", "FFC65911"),
        ("安保", "right", "FF2E75B6"),
    ]
    for name, halign, color in styles:
        dxf = DifferentialStyle(
            font=Font(color=color, bold=True),
            alignment=Alignment(horizontal=halign, vertical="center"),
        )
        rule = Rule(type="cellIs", operator="equal", formula=[f'"{name}"'], dxf=dxf)
        ws.conditional_formatting.add(k_range, rule)
